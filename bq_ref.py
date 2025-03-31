import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

def app():
    st.title("Reference Banque")
    
    uploaded_file = st.file_uploader("Choisissez votre fichier Excel", type=["xlsx"])

    if uploaded_file:
        try:
            # Read and prepare data
            # Read all sheets but we'll only process the first one
            xls = pd.ExcelFile(uploaded_file)
            df = pd.read_excel(xls, sheet_name=0, header=None)
            
            if df.shape[1] < 7:
                st.error("Le fichier doit contenir au moins 7 colonnes")
                return
                
            # Keep original values from column 5 (index 4) which will become RAW_REF
            original_ref_values = df.iloc[:, 4].values
            df.columns = ["DATE", "DROP", "RAW_LIB", "RAW_TIER", "RAW_REF", "DEBIT", "CREDIT"]
            df = df.drop(columns="DROP")
            
            # Convert to lowercase for matching and replace 'nan' with empty string
            # Force conversion to string type first to handle NaN/float values
            df["RAW_TIER"] = df["RAW_TIER"].fillna("").astype(str).str.lower()
            df["RAW_LIB"] = df["RAW_LIB"].fillna("").astype(str).str.lower()
            df["RAW_REF"] = df["RAW_REF"].fillna("").astype(str).str.lower()
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Processed"  # Rename the first sheet
            
            # Define dropdown options
            options = ["FELAH", "FAC", "FRAIS", "REMB", "PAIE", "COTIS", "IR", "RETENU MEDECIN", "RETENU AVOCAT"]
            
            # Create data validation
            dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
            ws.add_data_validation(dv)
                
            # Process each row with explicit row indexing to avoid blank rows
            row_idx = 1  # Start from row 2 (after headers)
            for idx, row in df.iterrows():
                # Convert to string and clean
                if pd.isna(original_ref_values[idx]):
                    original_ref = ""
                else:
                    original_ref = str(original_ref_values[idx]).strip()
                
                # Only auto-fill if the original cell was empty
                if original_ref == "":
                    # Safe string checking
                    raw_tier = str(row["RAW_TIER"]).lower() if pd.notna(row["RAW_TIER"]) else ""
                    raw_lib = str(row["RAW_LIB"]).lower() if pd.notna(row["RAW_LIB"]) else ""
                    raw_ref = str(row["RAW_REF"]).lower() if pd.notna(row["RAW_REF"]) else ""
                    
                    # Convert DEBIT and CREDIT to numeric values, defaulting to 0 if NaN
                    debit_val = float(row["DEBIT"]) if pd.notna(row["DEBIT"]) else 0
                    credit_val = float(row["CREDIT"]) if pd.notna(row["CREDIT"]) else 0
                    
                    if any(name in raw_tier for name in ["hamza", "youssef", "abdellah", "yousef", "rachid", "majdoubi", "touhami", "bikri", 
                                "bouzidi", "brahim", "derouach", "hatta", "benda", "khalid", "amine", "hamid", 
                                "mohammed", "mohamed", "sliman", "benbo", "dkhail", "charrad", "bouzid", "lehcen", 
                                "acha", "akil", "benz", "hassan", "hadri", "kharmo", "jilali", "houcin", "fellaha", 
                                "miloud", "asmaa", "bouchta", "ahmed", "afroukh", "znagui"]):
                        ref = "FELAH"
                    elif any(kw in raw_tier for kw in ["orange", "mamda", "onssa", "aswak", "brico", "carref", "wafabail", "cabinet", 
                                "trans", "redal", "refri", "secola", "dakar", "attijari", "temara", "kpa", "easy", 
                                "ajyad", "bioci", "must", "saidou", "bounmer", "pharmaci", "boutiq", "print", "moges", "fourni", "bois", 
                                "planex", "smurf", "asswak", "autor", "omnium", "dakkar", "idealalu", 
                                "deco new", "abcr", "azrou complexes", "da graph", "durofloor", "tissus", "electroplanet",
                                "forges de bazas", "globus", "environnement", "lab", "intra", "inwi",
                                "khadamat", "kitea", "trois", "lvs", "marjan", "ministre", "muneris", 
                                "consilia", "ola", "conseil", "pneumatique", "incendie", "sanitaire", 
                                "smpce", "star dec", "boulon"]):
                        ref = "FAC"
                    elif any(kw in raw_lib or kw in raw_tier for kw in ["frais", "commis", "pmeplus"]):
                        ref = "FRAIS"
                    elif raw_tier == "cnss":
                        ref = "COTIS"
                    elif "salaire" in raw_tier:
                        ref = "PAIE"
                    elif "cong" in raw_ref:
                        ref = "PAIE"
                    elif "ettoumy" in raw_tier:
                        ref = "PAIE"
                    elif "relanc" in raw_tier:
                        ref = "REMB"
                    elif raw_tier == "dgi" and debit_val > 50000:
                        ref = "IR"
                    elif "change" in raw_lib:
                        if debit_val > credit_val:
                            ref = "PERTE"
                        else:
                            ref = "GAIN"
                    else:
                        ref = ""  # Empty for dropdown
                else:
                    ref = original_ref  # Keep original value
                
                # Write row to Excel with explicit row index to avoid blank rows
                ws.cell(row=row_idx, column=1).value = row["DATE"]
                ws.cell(row=row_idx, column=2).value = row["RAW_LIB"]
                ws.cell(row=row_idx, column=3).value = row["RAW_TIER"]
                ws.cell(row=row_idx, column=4).value = ref
                ws.cell(row=row_idx, column=5).value = row["DEBIT"]
                ws.cell(row=row_idx, column=6).value = row["CREDIT"]
                
                # Add dropdown to empty cells or cells that had original values from dropdown options
                if ref == "" or (ref in options and ref == original_ref):
                    dv.add(ws[f"D{row_idx}"])  # D column is REF
                
                row_idx += 1  # Increment row counter
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add the second sheet from original file if it exists
            if len(xls.sheet_names) > 1:
                original_wb = load_workbook(uploaded_file)
                original_second_sheet = original_wb[xls.sheet_names[1]]
                new_sheet = wb.create_sheet(title=xls.sheet_names[1])
                
                for row in original_second_sheet.iter_rows():
                    new_sheet.append([cell.value for cell in row])
            
            # Save to buffer
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            st.success("Votre fichier est prêt !")
            st.download_button(
                "Télécharger le fichier",
                data=output,
                file_name="donnees_awb.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        except Exception as e:
            st.error(f"Erreur : {str(e)}")
            import traceback
            st.error(traceback.format_exc())  # Add this for detailed error info

if __name__ == "__main__":
    app()
