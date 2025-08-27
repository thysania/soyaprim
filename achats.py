import pandas as pd
import streamlit as st
from io import BytesIO
def app():
# Function to transform the data
    def transform_data(file):
    # Ensure the file is read properly as a BytesIO object
        df = pd.read_excel(file, header=None, engine="openpyxl")

    # Step 1: Delete unwanted columns (2, 3, 6, 7, 8 -> index 1, 2, 5, 6, 7)
        df.drop(columns=[1, 2, 5, 6, 7], inplace=True)

    # Step 2: Rename the remaining columns
        df.columns = ["Date", "PRODUIT", "TIERS", "TTC"]

    # Step 3: Add empty columns
        empty_columns = ["N FAC", "IF", "ICE", "MODE REGL", "DATE REGL", "TAUX TVA", "JOURNAL TRESORIE"]
        for col in empty_columns:
            df[col] = ""

    # Step 4: Add columns with default 0 value
        zero_columns = ["TVA", "CPT TVA"]
        for col in zero_columns:
            df[col] = 0

    # Step 5: Add "CPT HT" column with value 6111000000
        df["CPT HT"] = 6111000000

    # Step 6: Duplicate "TTC" column as "HT"
        df["HT"] = df["TTC"]

    # Step 7: Add "DESIGNATION" column as concatenation of "PRODUIT / TIERS" and delete "PRODUIT"
        df["DESIGNATION"] = df["PRODUIT"] + " / " + df["TIERS"]
        df.drop(columns=["PRODUIT"], inplace=True)

    # Step 8: Rearrange columns in the desired order and set data types
        columns_order = [
            "Date", "N FAC", "TIERS", "IF", "ICE", "DESIGNATION",
            "TTC", "HT", "TVA", "MODE REGL", "DATE REGL", "CPT HT",
            "CPT TVA", "TAUX TVA", "JOURNAL TRESORIE"
        ]
        df = df[columns_order]

    # Convert data types
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")  # Set invalid dates as NaT
        df["TIERS"] = df["TIERS"].astype(str) + "."
        df["DESIGNATION"] = df["DESIGNATION"].astype(str)
        df["TTC"] = df["TTC"].astype(float)
        df["HT"] = df["HT"].astype(float)
        df["CPT HT"] = df["CPT HT"].astype(int)
    
        return df
    
    # Streamlit app
    st.title("Achats")
    
    # File upload
    uploaded_file = st.file_uploader("Choisissez votre fichier Excel", type=["xlsx", "xls"])
    
    if uploaded_file:
        # Transform the uploaded file
        try:
            transformed_data = transform_data(uploaded_file)
    
            # Save the transformed data to a BytesIO object
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                transformed_data.to_excel(writer, index=False, sheet_name="Transformed Data")
                workbook = writer.book
                worksheet = writer.sheets["Transformed Data"]
    
                # Aooly formatting to the "Date" column
                from openpyxl.styles.numbers import FORMAT_DATE_DMYSLASH
                for col in worksheet.iter_cols(min_col=1, max_col=1, min_row=2, max_row=worksheet.max_row):
                    for cell in col:
                        cell.number_format = FORMAT_DATE_DMYSLASH #Set date format as dd/mm/yyyy
              
                # Style headers
                from openpyxl.styles import PatternFill, Font
                header_fill = PatternFill(start_color="4B9CD3", end_color="4B9CD3", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for cell in worksheet[1]:  # First row (headers)
                    cell.fill = header_fill
                    cell.font = header_font
    
            # Convert BytesIO to downloadable file
            output.seek(0)
            st.success("Votre fichier est prêt! 👍")
            st.download_button(
                label="Telecharger le fichier",
                data=output,
                file_name="Import_Achats.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"An error occurred: {e}")
if __name__ == "__main__":
    app()
