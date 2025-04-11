import streamlit as st
import pandas as pd
import numpy as np
import datetime as dt
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

def app():
    # Page title
    st.title("Banque")
    
    # File upload
    uploaded_file = st.file_uploader("Choisissez votre fichier Excel", type=["xlsx"])
    
    if uploaded_file is not None:
        try:
            # Read raw data (Sheet 1) and mappings (Sheet 2)
            raw_df = pd.read_excel(uploaded_file, sheet_name=0, header=None)
            mappings_df = pd.read_excel(uploaded_file, sheet_name=1, header=None)  # Assuming no headers in Sheet 2
    
            # Validate the number of columns in the raw file
            if raw_df.shape[1] != 7:
                st.error(f"8 colonnes attendues dans le fichier, mais {raw_df.shape[1]} trouvées. Veuillez vérifier la structure du fichier.")
            elif mappings_df.shape[1] < 2:
                st.error(f"Au moins 2 colonnes attendues dans la feuille de mappage, mais {mappings_df.shape[1]} trouvées. Veuillez vérifier la structure du fichier.")
            else:
                # Assign column names to raw data (7 columns)
                raw_df.columns = [
                    "DATE", "RAW_LIB", "RAW_TIER", "RAW_REF", 
                    "DEBIT", "CREDIT", "CA"
                ]
                
                # Convert numeric columns to proper numeric types from the beginning
                raw_df["DEBIT"] = pd.to_numeric(raw_df["DEBIT"], errors='coerce').fillna(0)
                raw_df["CREDIT"] = pd.to_numeric(raw_df["CREDIT"], errors='coerce').fillna(0)
                
                # Create a copy of rows to split - specifically for DGI with DEBIT=734
                rows_to_split = raw_df[(raw_df["RAW_TIER"].astype(str).str.contains("dgi", case=False)) & 
                                      (raw_df["DEBIT"] == 734)].copy()

                if not rows_to_split.empty:
                    # Create two new rows for each row to split
                    split_rows = []
                    indices_to_drop = []
                    
                    for idx, row in rows_to_split.iterrows():
                        indices_to_drop.append(idx)
                        
                        # First row - RETENU MEDECIN
                        medecin_row = row.copy()
                        medecin_row["RAW_REF"] = "RETENU MEDECIN"
                        medecin_row["DEBIT"] = 400
                        
                        # Second row - RETENU AVOCAT
                        avocat_row = row.copy()
                        avocat_row["RAW_REF"] = "RETENU AVOCAT"
                        avocat_row["DEBIT"] = 334
                        
                        split_rows.append(medecin_row)
                        split_rows.append(avocat_row)
                    
                    # Remove the original rows and add the new rows
                    raw_df = raw_df.drop(indices_to_drop)
                    
                    # Add the new rows
                    if split_rows:
                        split_df = pd.DataFrame(split_rows)
                        raw_df = pd.concat([raw_df, split_df], ignore_index=True)
    
                # Process TIERS (lookup RAW_TIER as wildcard in mappings sheet)
                def lookup_tiers(row):
                    raw_tier = row["RAW_TIER"]
                    debit = row["DEBIT"]
                    
                    if pd.isna(raw_tier):
                        return np.nan
                    
                    # Special case for "abdelatif saidou(medecin)"
                    if isinstance(raw_tier, str) and "saidou" in str(raw_tier).lower() and "medecin" in str(raw_tier).lower():
                        return "SAIDOU ABDELLATIPH"
                    
                    # Special case for "Wafabail" with specific DEBIT amounts
                    if isinstance(raw_tier, str) and "wafabail" in str(raw_tier).lower():
                        # Round to 2 decimal places to ensure exact matching
                        debit_rounded = round(debit, 2)
                        
                        if debit_rounded == 28780.38:
                            return "WAFABAIL CONTRAT S0514740 SMARTEC"
                        elif debit_rounded == 13790.77:
                            return "WAFABAIL CONTRAT S0623820 FORGEZ DE BASAS"
                        elif debit_rounded == 10204.08:
                            return "WAFABAIL CONTRAT S0514750 MAP MAGHREB"
                        elif debit_rounded == 20408.17 or debit_rounded == 20139.84:
                            return "WAFABAIL CONTRAT S0514770 MAP MAGHREB"
                    
                    # Find rows in mappings_df where the pattern matches RAW_TIER (wildcard)
                    matches = mappings_df[mappings_df[0].astype(str).str.contains(str(raw_tier), case=False, na=False)]
                    if not matches.empty:
                        return matches.iloc[0, 1]  # Return the first matching value from column 1
                    return np.nan
    
                # Apply the lookup_tiers function to each row
                raw_df["TIERS"] = raw_df.apply(lookup_tiers, axis=1)
    
                # Process CPT (apply conditions)
                # Ensure all conditions are boolean arrays
                conditions = [
                    # Rule 0: Column CA equals 1
                    (raw_df["CA"] == 1),
                    # Rule 1: RAW_TIER starts with "FRUL" (case-insensitive)
                    raw_df["RAW_TIER"].astype(str).str.upper().str.startswith("FRUL", na=False),
                    # Rule 2: RAW_FILTER starts with "SALAIRE" or NAT == "PAIE"
                    (raw_df["RAW_REF"].astype(str).str.upper().str.startswith("CONGÉ", na=False)) | (raw_df["RAW_REF"] == "PAIE"),
                    # Rule 3: RAW_TIER == "CNSS" or NAT == "COTIS"
                    (raw_df["RAW_TIER"].astype(str).str.upper() == "CNSS") | (raw_df["RAW_REF"] == "COTIS"),
                    # Rule 4: RAW_REF == FRAIS
                    (raw_df["RAW_REF"] == "FRAIS") | raw_df["RAW_LIB"].astype(str).str.upper().str.startswith("FRAIS", na=False),
                    # Rule 5: RAW_REF equals PERTE
                    (raw_df["RAW_REF"] == "PERTE"),
                    # Rule 5.1: RAW_REF equals GAIN
                    (raw_df["RAW_REF"] == "GAIN"),
                    # Rule 6: NAT == "FELAH" or TIERS contains specified strings
                    (raw_df["RAW_REF"] == "FELAH") | 
                    raw_df["TIERS"].astype(str).str.contains("|".join([
                        "AJYAD", "ASWAK", "ATTIJARI", "AUTOROUTE", "BIOCI", "BOIS", "BOUNMER", 
                        "BRICO", "CABINET", "CARREF", "consilia", "conseil", "da graph", "deco new", 
                        "durofloor", "electroplanet", "environnement", "FOURNI", "BAZAS", 
                        "globus", "hamri tissus", "ideal", "inwi", "incendie", "khadamat", "kitea", 
                        "KPA", "lab", "lvs", "MAMDA", "MARAHIL", "marjan", "ministre", "MOGES", "MUST", "ORANGE", 
                        "PLANEX", "pneumatique", "PRINT", "REDAL", "relanc", "REFRI", "SANITAIRE", 
                        "secola", "SMURF", "smpce", "SOLUTIONS", "STARDEC", "TEMARA", "trois", "WAFABAIL", 
                        "intra", "abcr", "boulon", "bmj", "Khadamat", "TRANS", "AFROUKH", "onssa", "ZIMBA", 
                        "STORES", "GOLD", "COMPLEX", "LEGNO", "TISSUS", "DAKAR", "PROJET", "D.A"
                    ]), case=False, na=False),
                    # Rule 7: RAW_REF == "IR"
                    (raw_df["RAW_REF"] == "IR"),
                    # Rule 8: RAW_REF == "RETENU MEDECIN"
                    (raw_df["RAW_REF"] == "RETENU MEDECIN"),
                    # Rule 9: RAW_REF == "RETENU AVOCAT"
                    (raw_df["RAW_REF"] == "RETENU AVOCAT"),
                    # Rule 10: RAW_TIER contains "SAIDOU"
                    (raw_df["RAW_TIER"].astype(str).str.upper().str.contains("SAIDOU", na=False))
                    # Rule 11: RAW_TIER contains "vignette"
                    (raw_df["RAW_TIER"].astype(str).str.upper().str.contains("SVIGNETTE", na=False))
                ]
    
                choices = [
                    3497000000,  # Rule 0
                    3421000000,  # Rule 1
                    4432000000,  # Rule 2
                    4441000000,  # Rule 3
                    6147300000,  # Rule 4
                    6331000000,  # Rule 5
                    7331000000,  # Rule 5.1
                    4411000000,  # Rule 6
                    4452500000,  # Rule 7
                    4458110100,  # Rule 8
                    4458110200,  # Rule 9
                    4411000000,  # Rule 10
                    4452110000   # Rule 11
                ]
    
                # Use float type for CPT to support NaN values
                raw_df["CPT"] = np.select(conditions, choices, default=np.nan).astype(float)
    
                # Process LIB (concatenate RAW_LIB/NAT/RAW_TIER) - ignoring empty cells
                raw_df["LIB"] = raw_df.apply(
                    lambda row: " / ".join(
                        filter(None, [
                            str(row["RAW_LIB"]).strip() if pd.notna(row["RAW_LIB"]) else None,
                            str(row["RAW_REF"]).strip() if pd.notna(row["RAW_REF"]) else None,
                            str(row["TIERS"]).strip() if pd.notna(row["TIERS"]) else (
                                str(row["RAW_TIER"]).strip() if pd.notna(row["RAW_TIER"]) else None
                            )
                        ])
                    ),
                    axis=1
                )
                
                # Create final DataFrame with specified columns and formatting
                result_df = pd.DataFrame({
                    "DATE": pd.to_datetime(raw_df["DATE"]).dt.strftime("%d/%m/%Y"),
                    "N PIECE": np.nan,
                    "CPT": raw_df["CPT"],
                    "TIERS": raw_df["TIERS"],
                    "LIB": raw_df["LIB"].astype(str),
                    "REF": np.nan,
                    "DEBIT": raw_df["DEBIT"].round(2),
                    "CREDIT": raw_df["CREDIT"].round(2)
                })
    
                # Replace "nan" strings with truly empty cells
                result_df["CPT"] = result_df["CPT"].replace("nan", np.nan)
                result_df["TIERS"] = result_df["TIERS"].replace("nan", np.nan)
                
                # Sort the dataframe by DATE from old to new
                # First convert back to datetime for proper sorting
                result_df["DATE_SORT"] = pd.to_datetime(result_df["DATE"], format="%d/%m/%Y")
                result_df = result_df.sort_values(by="DATE_SORT", ascending=True)
                result_df = result_df.drop(columns=["DATE_SORT"])  # Remove the sorting column
    
                # Show preview of the result
                st.write("### Aperçu des Données Transformées")
                st.dataframe(result_df.head())
                
                ###############################################
                # PIVOT FOR CPT DETAILS SHEET
                ###############################################
                # Create a flattened version of the data for the pivot table
                result_df_for_pivot = result_df.copy()
                
                # Ensure DEBIT and CREDIT are numeric values
                result_df_for_pivot["DEBIT"] = pd.to_numeric(result_df_for_pivot["DEBIT"], errors='coerce').fillna(0)
                result_df_for_pivot["CREDIT"] = pd.to_numeric(result_df_for_pivot["CREDIT"], errors='coerce').fillna(0)
                
                # Replace NaN in CPT with "Vide" string
                result_df_for_pivot["CPT"] = result_df_for_pivot["CPT"].fillna("Vide")
                
                # Group by CPT and LIB to get sums
                grouped_data = result_df_for_pivot.groupby(["CPT", "LIB"]).agg({
                    "DEBIT": "sum",
                    "CREDIT": "sum"
                }).reset_index()
                
                # Add CPT subtotals
                cpt_totals = result_df_for_pivot.groupby("CPT").agg({
                    "DEBIT": "sum",
                    "CREDIT": "sum"
                }).reset_index()
                
                # Add "Total" to LIB column for the CPT subtotals
                cpt_totals["LIB"] = "Total"
                
                # Add grand total
                grand_total = pd.DataFrame({
                    "CPT": ["Total"],
                    "LIB": [""],
                    "DEBIT": [result_df_for_pivot["DEBIT"].sum()],
                    "CREDIT": [result_df_for_pivot["CREDIT"].sum()]
                })
                
                # Combine all data - First grouped data, then CPT totals, then grand total
                combined_data = pd.concat([grouped_data, cpt_totals, grand_total], ignore_index=True)
                
                # Sort the dataframe - this keeps the rows with the same CPT together
                # and ensures the "Vide" rows are together as well
                combined_data = combined_data.sort_values(by=["CPT", "LIB"])
                
                # Move Total CPT to the end
                total_row = combined_data[combined_data["CPT"] == "Total"]
                combined_data = combined_data[combined_data["CPT"] != "Total"]
                combined_data = pd.concat([combined_data, total_row], ignore_index=True)
                
                ###############################################
                # PIVOT FOR EMPTY TIERS SHEET
                ###############################################
                # Create a copy for the TIERS pivot
                tiers_df = result_df.copy()
                
                # Ensure DEBIT and CREDIT are numeric
                tiers_df["DEBIT"] = pd.to_numeric(tiers_df["DEBIT"], errors='coerce').fillna(0)
                tiers_df["CREDIT"] = pd.to_numeric(tiers_df["CREDIT"], errors='coerce').fillna(0)
                
                # Filter for empty TIERS
                empty_tiers_df = tiers_df[tiers_df["TIERS"].isna()]
                
                # Replace NaN in TIERS with "Vide" for grouping
                empty_tiers_df["TIERS"] = "TIERS VIDE"
                
                # Group by TIERS and LIB to get sums
                tiers_grouped = empty_tiers_df.groupby(["TIERS", "LIB"]).agg({
                    "DEBIT": "sum",
                    "CREDIT": "sum"
                }).reset_index()
                
                # Add TIERS subtotal
                tiers_total = empty_tiers_df.groupby("TIERS").agg({
                    "DEBIT": "sum",
                    "CREDIT": "sum"
                }).reset_index()
                tiers_total["LIB"] = "Total"
                
                # Combine the data
                tiers_combined = pd.concat([tiers_grouped, tiers_total], ignore_index=True)
                
                # Sort to ensure total appears at the end
                tiers_combined = tiers_combined.sort_values(by=["TIERS", "LIB"])
                tiers_combined_total = tiers_combined[tiers_combined["LIB"] == "Total"]
                tiers_combined = tiers_combined[tiers_combined["LIB"] != "Total"]
                tiers_combined = pd.concat([tiers_combined, tiers_combined_total], ignore_index=True)
                
                ###############################################
                # CREATE EXCEL FILE WITH ALL SHEETS
                ###############################################
                # Create Excel file for download
                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    # Write first sheet - main data
                    result_df.to_excel(writer, index=False, sheet_name="Données Transformées", na_rep="")
                    
                    # Write second sheet - CPT pivot data
                    combined_data.to_excel(writer, sheet_name="Détails Comptes", index=False)
                    
                    # Write third sheet - TIERS pivot data (only empty TIERS)
                    tiers_combined.to_excel(writer, sheet_name="Détails Tiers", index=False)
                    
                    # Get the worksheets to apply formatting
                    workbook = writer.book
                    main_sheet = writer.sheets["Données Transformées"]
                    pivot_sheet = writer.sheets["Détails Comptes"]
                    tiers_sheet = writer.sheets["Détails Tiers"]
                    
                    ###############################################
                    # FORMAT MAIN SHEET
                    ###############################################
                    # Auto-fit columns
                    for col_num, column in enumerate(result_df.columns, 1):
                        column_width = max(
                            result_df[column].astype(str).map(len).max(),  # Width of the data
                            len(str(column))  # Width of the header
                        ) + 2  # Add a little extra space
                        main_sheet.column_dimensions[get_column_letter(col_num)].width = column_width
                    
                    # Format headers with color #2596be
                    for cell in main_sheet[1]:
                        cell.fill = PatternFill(start_color="2596BE", end_color="2596BE", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")  # White text for better contrast
                    
                    ###############################################
                    # FORMAT CPT PIVOT SHEET
                    ###############################################
                    # Format headers with the same color
                    for cell in pivot_sheet[1]:
                        cell.fill = PatternFill(start_color="2596BE", end_color="2596BE", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    
                    # Auto-fit columns for pivot table
                    for col_num, column in enumerate(combined_data.columns, 1):
                        if col_num == 2:  # LIB column
                            pivot_sheet.column_dimensions[get_column_letter(col_num)].width = 60
                        else:
                            column_width = max(
                                combined_data[column].astype(str).map(len).max(),
                                len(str(column))
                            ) + 2
                            pivot_sheet.column_dimensions[get_column_letter(col_num)].width = column_width
                    
                    # Define color fills
                    light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    total_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    
                    # Apply formatting to pivot sheet rows
                    for row_idx in range(2, pivot_sheet.max_row + 1):  # Start from row 2 (after header)
                        cpt_value = pivot_sheet.cell(row=row_idx, column=1).value
                        lib_value = pivot_sheet.cell(row=row_idx, column=2).value
                        
                        # Color all rows where CPT is "Vide" with light red
                        if cpt_value == "Vide":
                            for col_idx in range(1, pivot_sheet.max_column + 1):
                                pivot_sheet.cell(row=row_idx, column=col_idx).fill = light_red_fill
                        
                        # Format total rows
                        if (cpt_value == "Total") or (lib_value == "Total"):
                            for col_idx in range(1, pivot_sheet.max_column + 1):
                                cell = pivot_sheet.cell(row=row_idx, column=col_idx)
                                cell.fill = total_fill
                                cell.font = Font(bold=True)
                    
                    ###############################################
                    # FORMAT TIERS PIVOT SHEET
                    ###############################################
                    # Format headers with the same color
                    for cell in tiers_sheet[1]:
                        cell.fill = PatternFill(start_color="2596BE", end_color="2596BE", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    
                    # Auto-fit columns for TIERS pivot table
                    for col_num, column in enumerate(tiers_combined.columns, 1):
                        if col_num == 2:  # LIB column
                            tiers_sheet.column_dimensions[get_column_letter(col_num)].width = 60
                        else:
                            column_width = max(
                                tiers_combined[column].astype(str).map(len).max(),
                                len(str(column))
                            ) + 2
                            tiers_sheet.column_dimensions[get_column_letter(col_num)].width = column_width
                    
                    # Apply formatting to TIERS pivot sheet rows
                    for row_idx in range(2, tiers_sheet.max_row + 1):  # Start from row 2 (after header)
                        lib_value = tiers_sheet.cell(row=row_idx, column=2).value
                        
                        # Format total rows
                        if lib_value == "Total":
                            for col_idx in range(1, tiers_sheet.max_column + 1):
                                cell = tiers_sheet.cell(row=row_idx, column=col_idx)
                                cell.fill = total_fill
                                cell.font = Font(bold=True)
                
                output.seek(0)
                st.success("Votre fichier est prêt !")
                st.download_button(
                    label="Télécharger le fichier",
                    data=output,
                    file_name="import_awb.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        except Exception as e:
            st.error(f"Une erreur s'est produite : {e}")
            # For debugging
            st.error(f"Détails: {str(e)}")
            import traceback
            st.error(traceback.format_exc())
            
if __name__ == "__main__":
    app()
